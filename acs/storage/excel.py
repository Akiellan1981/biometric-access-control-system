"""On-demand Excel export (spec FR-4.2 / AC-5).

Columns: Name, Date, Time, Method, Result, Score.
"""
from __future__ import annotations

from pathlib import Path


def export_events(rows, out_path: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["Name", "Date", "Time", "Method", "Result", "Score"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ts = r["ts"] or ""
        date, _, tm = ts.partition(" ")
        ws.append([
            r["name"], date, tm, r["method"], r["result"], round(r["score"] or 0, 4),
        ])

    for col, width in zip("ABCDEF", (22, 12, 10, 12, 16, 8)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
