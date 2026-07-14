from acs.storage.excel import export_events


def test_export(tmp_path):
    rows = [
        {"name": "Asha", "ts": "2026-06-24 09:00:00", "method": "face",
         "result": "granted", "score": 0.712},
        {"name": "unknown", "ts": "2026-06-24 09:05:00", "method": "face",
         "result": "denied-spoof", "score": 0.1},
    ]
    out = export_events(rows, tmp_path / "a.xlsx")
    assert out.exists()

    from openpyxl import load_workbook
    ws = load_workbook(out).active
    assert [c.value for c in ws[1]] == ["Name", "Date", "Time", "Method", "Result", "Score"]
    assert ws["A2"].value == "Asha"
    assert ws["B2"].value == "2026-06-24"
    assert ws["C2"].value == "09:00:00"
